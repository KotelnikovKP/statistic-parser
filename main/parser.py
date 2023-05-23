from enum import Enum

import openpyxl


class ParserStatus(Enum):
    VALID = 'valid'
    INVALID_FILE_TYPE = 'Неверный тип файла.'
    INVALID_STRUCTURE = 'Неверная структура данных.'
    NO_DATA = 'Нет данных.'


class DataFileParser:
    def __init__(self, file):
        self._work_sheet = None
        self._valid = False
        try:
            self._work_sheet = openpyxl.load_workbook(file).active
            self._parse_header()
        except Exception as e:
            self._status = ParserStatus.INVALID_FILE_TYPE.value

    def _parse_header(self):
        """
            Функция парсинга заголовка (здесь же определяется валидность входного файла)
        """
        # Проверить на валидность структуры заголовка файла (наличие первых двух колонок 'id' и 'company'
        # и наличие первых трех строчек с индикаторами)
        if self._work_sheet.max_row < 3 or self._work_sheet.max_column < 2 or \
                self._work_sheet.cell(row=1, column=1).value != 'id' or \
                self._work_sheet.cell(row=1, column=2).value != 'company' or \
                self._work_sheet.cell(row=2, column=1).value is not None or \
                self._work_sheet.cell(row=2, column=2).value is not None or \
                self._work_sheet.cell(row=3, column=1).value is not None or \
                self._work_sheet.cell(row=3, column=2).value is not None:
            self._status = ParserStatus.INVALID_STRUCTURE.value
            return

        # Проверить наличие данных в файле (индикаторы и значения)
        if self._work_sheet.max_row < 4 or self._work_sheet.max_column < 3:
            self._status = ParserStatus.NO_DATA.value
            return

        self._status = ParserStatus.VALID.value
        self._valid = True

        # Определить для каждой колонки трех индикаторов (1-го, 2-го и 3-го уровней, каждый из своей строчки)
        self._head_indicators = list()
        head_indicator_1 = head_indicator_2 = None
        for col in self._work_sheet.iter_cols(3, self._work_sheet.max_column):
            if col[0].value:
                head_indicator_1 = col[0].value
            if col[1].value:
                head_indicator_2 = col[1].value
            if col[2].value:
                self._head_indicators.append((col[0].column, head_indicator_1, head_indicator_2, col[2].value))

    @property
    def status(self):
        return self._status

    @property
    def valid(self):
        return self._valid

    def __str__(self):
        if self._valid:
            result = ''
            for row in range(0, self._work_sheet.max_row):
                for col in self._work_sheet.iter_cols(1, self._work_sheet.max_column):
                    result += f"{col[row].value}\t"
                result += '\n'
            return result
        else:
            return self._status

    def __iter__(self):
        """
            Определение объекта класса как итератора (задание начальных значений индексов)
        """
        self._row = 4
        self._col = 0
        return self

    def __next__(self):
        """
            Функция обработчик итерации (возвращает словарь с единичным значением,
            соответствующее набору параметров (id, company, индикатор 1-го уровня,
            индикатор 2-го уровня, индикатор 3-го уровня)
        """
        if self._status != ParserStatus.VALID.value or self._row > self._work_sheet.max_row:
            raise StopIteration
        result = {
            'id': self._work_sheet.cell(row=self._row, column=1).value,
            'company': self._work_sheet.cell(row=self._row, column=2).value,
            'first_level_indicator': self._head_indicators[self._col][1],
            'second_level_indicator': self._head_indicators[self._col][2],
            'third_level_indicator': self._head_indicators[self._col][3],
            'value': self._work_sheet.cell(row=self._row, column=self._head_indicators[self._col][0]).value
        }
        self._col += 1
        if self._col >= len(self._head_indicators):
            self._col = 0
            self._row += 1
        return result
